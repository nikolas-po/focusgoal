"""Сервис экспорта данных (ТЗ FR-011)"""
import json
import csv
from pathlib import Path
from datetime import datetime
from sqlalchemy.orm import Session
from src.config.settings import Settings


class ExportService:
    def __init__(self, db: Session, user_id: int):
        self.db = db
        self.user_id = user_id
        self.settings = Settings()

    def export_to_json(self, output_path: str, encrypt: bool = True) -> str:
        from src.models.goal import Goal
        from src.models.habit import Habit
        from src.models.focus_session import FocusSession
        from src.models.completion_log import CompletionLog

        data = {
            "export_date": datetime.now().isoformat(),
            "app": "FocusGoal",
            "version": "1.0.0",
            "user_id": self.user_id,
            "goals": [],
            "habits": [],
            "focus_sessions": [],
            "completion_logs": [],
        }

        for g in self.db.query(Goal).filter(Goal.user_id == self.user_id).all():
            data["goals"].append({
                "id": g.id, "name": g.name, "description": g.description,
                "deadline": g.deadline.isoformat() if g.deadline else None,
                "priority_id": g.priority_id, "repeat_type_id": g.repeat_type_id,
                "fail_behavior_id": g.fail_behavior_id, "status_id": g.status_id,
                "created_at": g.created_at.isoformat(),
            })

        for h in self.db.query(Habit).filter(Habit.user_id == self.user_id).all():
            data["habits"].append({
                "id": h.id, "name": h.name, "description": h.description,
                "type_id": h.type_id, "mode_id": h.mode_id,
                "target_value": h.target_value, "current_streak": h.current_streak,
                "max_streak": h.max_streak, "status_id": h.status_id,
                "start_date": h.start_date.isoformat() if h.start_date else None,
            })

        for s in self.db.query(FocusSession).filter(FocusSession.user_id == self.user_id).all():
            data["focus_sessions"].append({
                "id": s.id, "start_time": s.start_time.isoformat(),
                "planned_duration": s.planned_duration,
                "actual_duration": s.actual_duration, "status_id": s.status_id,
            })

        for l in self.db.query(CompletionLog).filter(CompletionLog.user_id == self.user_id).all():
            data["completion_logs"].append({
                "id": l.id, "object_type_id": l.object_type_id, "object_id": l.object_id,
                "completed_at": l.completed_at.isoformat(),
                "progress": l.progress, "comment": l.comment,
            })

        json_bytes = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
        output_file = Path(output_path)

        if encrypt and self.settings.ENCRYPTION_KEY:
            from src.services.encryption_service import EncryptionService
            crypto = EncryptionService(self.settings.ENCRYPTION_KEY)
            encrypted = crypto.encrypt(json_bytes)
            output_file.write_bytes(b"FGENC" + crypto.salt + encrypted)
        else:
            output_file.write_bytes(json_bytes)

        return str(output_file)

    def export_to_csv(self, output_path: str) -> str:
        from src.models.goal import Goal
        from src.models.habit import Habit
        from src.models.focus_session import FocusSession

        goals = self.db.query(Goal).filter(Goal.user_id == self.user_id).all()
        habits = self.db.query(Habit).filter(Habit.user_id == self.user_id).all()
        sessions = self.db.query(FocusSession).filter(FocusSession.user_id == self.user_id).all()

        prio = {1: "Высокий", 2: "Средний", 3: "Низкий"}
        gstat = {1: "Активна", 2: "Выполнена", 3: "Просрочена", 4: "Отменена", 5: "Удалена", 6: "В архиве"}
        htype = {1: "Ежедневная", 2: "Еженедельная", 3: "Ежемесячная"}
        hmode = {1: "Бинарная", 2: "Количественная"}
        hstat = {1: "Активна", 2: "В архиве", 3: "Отключена", 4: "Удалена"}
        fstat = {1: "Завершена", 2: "Прервана", 3: "Прервана внешне"}

        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["# Отчёт FocusGoal", datetime.now().strftime("%d.%m.%Y %H:%M")])
            writer.writerow([])
            writer.writerow(["=== ЦЕЛИ ==="])
            writer.writerow(["ID", "Название", "Описание", "Срок", "Приоритет", "Статус", "Создана"])
            for g in goals:
                writer.writerow([g.id, g.name, g.description or "",
                                  g.deadline.strftime("%d.%m.%Y") if g.deadline else "",
                                  prio.get(g.priority_id, ""), gstat.get(g.status_id, ""),
                                  g.created_at.strftime("%d.%m.%Y")])
            writer.writerow([])
            writer.writerow(["=== ПРИВЫЧКИ ==="])
            writer.writerow(["ID", "Название", "Тип", "Режим", "Серия", "Макс.серия", "Статус"])
            for h in habits:
                writer.writerow([h.id, h.name, htype.get(h.type_id, ""),
                                  hmode.get(h.mode_id, ""), h.current_streak, h.max_streak,
                                  hstat.get(h.status_id, "")])
            writer.writerow([])
            writer.writerow(["=== ФОКУС-СЕССИИ ==="])
            writer.writerow(["ID", "Начало", "Плановая (мин)", "Факт. (мин)", "Статус"])
            for s in sessions:
                writer.writerow([s.id, s.start_time.strftime("%d.%m.%Y %H:%M"),
                                  s.planned_duration, s.actual_duration or "",
                                  fstat.get(s.status_id, "")])

        return output_path


    def export_to_xlsx(self, output_path: str) -> str:
        """Экспорт в Excel xlsx (ТЗ FR-011, openpyxl)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise RuntimeError("openpyxl не установлен: pip install openpyxl")

        from src.models.goal import Goal
        from src.models.habit import Habit
        from src.models.focus_session import FocusSession

        wb = openpyxl.Workbook()

        #  Цели 
        ws = wb.active
        ws.title = "Цели"
        hfill = PatternFill("solid", fgColor="4CAF50")
        hfont = Font(bold=True, color="FFFFFF")
        for ci, h in enumerate(["ID","Название","Описание","Срок","Приоритет","Статус","Создана"], 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hfill; c.font = hfont
            c.alignment = Alignment(horizontal="center")
        prio  = {1:"Высокий",2:"Средний",3:"Низкий"}
        gstat = {1:"Активна",2:"Выполнена",3:"Просрочена",4:"Отменена",5:"Удалена",6:"В архиве"}
        goals = self.db.query(Goal).filter(Goal.user_id == self.user_id).all()
        for ri, g in enumerate(goals, 2):
            ws.cell(ri,1,g.id); ws.cell(ri,2,g.name)
            ws.cell(ri,3,g.description or "")
            ws.cell(ri,4,g.deadline.strftime("%d.%m.%Y") if g.deadline else "")
            ws.cell(ri,5,prio.get(g.priority_id,"")); ws.cell(ri,6,gstat.get(g.status_id,""))
            ws.cell(ri,7,g.created_at.strftime("%d.%m.%Y"))

        #  Привычки 
        ws2 = wb.create_sheet("Привычки")
        h2fill = PatternFill("solid", fgColor="2196F3")
        for ci, h in enumerate(["ID","Название","Тип","Режим","Серия","Макс.серия","Статус"], 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.fill = h2fill; c.font = Font(bold=True, color="FFFFFF")
        htype = {1:"Ежедневная",2:"Еженедельная",3:"Ежемесячная"}
        hmode = {1:"Бинарная",2:"Количественная"}
        hstat = {1:"Активна",2:"В архиве",3:"Отключена",4:"Удалена"}
        habits = self.db.query(Habit).filter(Habit.user_id == self.user_id).all()
        for ri, h in enumerate(habits, 2):
            ws2.cell(ri,1,h.id); ws2.cell(ri,2,h.name)
            ws2.cell(ri,3,htype.get(h.type_id,"")); ws2.cell(ri,4,hmode.get(h.mode_id,""))
            ws2.cell(ri,5,h.current_streak); ws2.cell(ri,6,h.max_streak)
            ws2.cell(ri,7,hstat.get(h.status_id,""))

        #  Фокус-сессии 
        ws3 = wb.create_sheet("Фокус-сессии")
        h3fill = PatternFill("solid", fgColor="FF9800")
        for ci, h in enumerate(["ID","Начало","Плановая (мин)","Факт. (мин)","Статус"], 1):
            c = ws3.cell(row=1, column=ci, value=h)
            c.fill = h3fill; c.font = Font(bold=True, color="FFFFFF")
        fstat = {1:"Завершена",2:"Прервана",3:"Прервана внешне"}
        sessions = self.db.query(FocusSession).filter(FocusSession.user_id == self.user_id).all()
        for ri, s in enumerate(sessions, 2):
            ws3.cell(ri,1,s.id)
            ws3.cell(ri,2,s.start_time.strftime("%d.%m.%Y %H:%M"))
            ws3.cell(ri,3,s.planned_duration)
            ws3.cell(ri,4,s.actual_duration or "")
            ws3.cell(ri,5,fstat.get(s.status_id,""))

        wb.save(output_path)
        return output_path

    def export_to_pdf(self, output_path: str) -> str:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                             Table, TableStyle)
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            raise RuntimeError("reportlab не установлен: pip install reportlab")

        # Регистрация шрифта с поддержкой кириллицы
        _font_name = "DejaVuSans"
        _font_bold = "DejaVuSans-Bold"
        _registered = False
        for base_dir in (
            "/usr/share/fonts/truetype/dejavu",
            "/usr/share/fonts/dejavu",
            "/usr/share/fonts/TTF",
            "/usr/local/share/fonts",
            str(Path.home() / ".local" / "share" / "fonts"),
        ):
            import os
            rp = os.path.join(base_dir, "DejaVuSans.ttf")
            rb = os.path.join(base_dir, "DejaVuSans-Bold.ttf")
            if os.path.exists(rp):
                try:
                    pdfmetrics.registerFont(TTFont(_font_name, rp))
                    if os.path.exists(rb):
                        pdfmetrics.registerFont(TTFont(_font_bold, rb))
                    else:
                        _font_bold = _font_name
                    _registered = True
                    break
                except Exception:
                    pass
        if not _registered:
            try:
                import matplotlib.font_manager as fm
                font_path = Path(fm.findfont("DejaVu Sans", fallback_to_default=False))
                if font_path.exists():
                    pdfmetrics.registerFont(TTFont(_font_name, str(font_path)))
                    _registered = True
            except Exception:
                pass
        if not _registered:
            _font_name = "Helvetica"
            _font_bold = "Helvetica-Bold"

        from src.models.goal import Goal
        from src.models.habit import Habit
        from src.models.focus_session import FocusSession

        doc = SimpleDocTemplate(output_path, pagesize=A4,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        green = colors.HexColor("#4CAF50")
        story = []

        styles["Normal"].fontName = _font_name
        styles["Heading2"].fontName = _font_bold
        title_style = ParagraphStyle("Title2", parent=styles["Title"],
                                     fontSize=20, textColor=green, spaceAfter=6,
                                     fontName=_font_bold)
        story.append(Paragraph("Отчёт FocusGoal", title_style))
        story.append(Paragraph(
            f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            styles["Normal"]))
        story.append(Spacer(1, 0.5*cm))

        goals = self.db.query(Goal).filter(Goal.user_id == self.user_id).all()
        habits = self.db.query(Habit).filter(Habit.user_id == self.user_id).all()
        sessions = self.db.query(FocusSession).filter(FocusSession.user_id == self.user_id).all()

        completed_goals = sum(1 for g in goals if g.status_id == 2)
        focus_total = sum((s.actual_duration or 0) for s in sessions if s.status_id == 1)

        story.append(Paragraph("Общая статистика", styles["Heading2"]))
        summary = [
            ["Показатель", "Значение"],
            ["Всего целей", str(len(goals))],
            ["Выполнено целей", str(completed_goals)],
            ["% выполнения", f"{completed_goals/len(goals)*100:.0f}%" if goals else "0%"],
            ["Всего привычек", str(len(habits))],
            ["Всего фокус-сессий", str(len(sessions))],
            ["Суммарное время в фокусе", f"{focus_total // 60}ч {focus_total % 60}мин"],
        ]
        t = Table(summary, colWidths=[9*cm, 6*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), green),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), _font_bold),
            ("FONTNAME", (0, 1), (-1, -1), _font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.5*cm))

        if goals:
            story.append(Paragraph("Список целей", styles["Heading2"]))
            prio = {1: "Высокий", 2: "Средний", 3: "Низкий"}
            gstat = {1: "Активна", 2: "Выполнена", 3: "Просрочена",
                     4: "Отменена", 5: "Удалена", 6: "В архиве"}
            goal_data = [["Название", "Срок", "Приоритет", "Статус"]]
            for g in goals[:25]:
                goal_data.append([
                    g.name[:40],
                    g.deadline.strftime("%d.%m.%Y") if g.deadline else "—",
                    prio.get(g.priority_id, ""),
                    gstat.get(g.status_id, ""),
                ])
            gt = Table(goal_data, colWidths=[7*cm, 3*cm, 3*cm, 3*cm])
            gt.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2196F3")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), _font_bold),
                ("FONTNAME", (0, 1), (-1, -1), _font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(gt)

        doc.build(story)
        return output_path
